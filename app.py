import streamlit as st
import pandas as pd
import os
import tempfile
from math import ceil
import zipfile
from io import BytesIO
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

# Configurar a página
st.set_page_config(
    page_title="Andrezin - Ferramentas de Planilhas",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS personalizado FINAL
st.markdown("""
<style>
    /* Fundo cinza permanente */
    .stApp {
        background-color: #f8f9fa !important;
    }
    
    /* Forçar tema claro */
    [data-testid="stAppViewContainer"] {
        background-color: #f8f9fa !important;
    }
    
    /* Remover elementos do Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stDeployButton {visibility: hidden;}
    
    /* Container principal */
    .main-container {
        background: white;
        border-radius: 20px;
        padding: 2rem;
        margin: 1rem auto;
        box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        border: 1px solid #e9ecef;
        max-width: 1200px;
    }
    
    .main-header {
        font-size: 2.5rem;
        text-align: center;
        margin-bottom: 0.5rem;
        font-weight: 800;
        background: linear-gradient(135deg, #7D3C98 0%, #8E44AD 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .subtitle {
        text-align: center;
        color: #6c757d;
        font-size: 1.1rem;
        margin-bottom: 2rem;
        font-weight: 300;
    }
    
    /* UPLOADER COM FUNDO ROXO E LETRAS BRANCAS */
    .upload-section {
        background: linear-gradient(135deg, #7D3C98 0%, #8E44AD 100%);
        border-radius: 15px;
        padding: 2rem;
        margin: 1rem 0;
        color: white !important;
    }
    
    .upload-section h3 {
        color: white !important;
        margin-bottom: 1rem;
    }
    
    .upload-section .stFileUploader > label {
        color: white !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
    }
    
    .upload-section .stFileUploader > section > div {
        background-color: rgba(255, 255, 255, 0.1) !important;
        border: 2px dashed rgba(255, 255, 255, 0.5) !important;
        border-radius: 15px !important;
        color: white !important;
    }
    
    .upload-section .stFileUploader > section > div:hover {
        background-color: rgba(255, 255, 255, 0.2) !important;
        border-color: rgba(255, 255, 255, 0.8) !important;
    }
    
    .upload-section .stFileUploader > section > div > div > small {
        color: rgba(255, 255, 255, 0.9) !important;
        font-size: 0.8rem !important;
    }
    
    .upload-section .stFileUploader > section > div > div > button {
        background: rgba(255, 255, 255, 0.9) !important;
        color: #7D3C98 !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.4rem 1.2rem !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
    }
    
    .upload-section .stFileUploader > section > div > div > button:hover {
        background: white !important;
        color: #7D3C98 !important;
    }
    
    /* Cards modernos */
    .feature-card {
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
        border-radius: 15px;
        padding: 2rem 1.5rem;
        margin: 1rem;
        border: 2px solid #e9ecef;
        text-align: center;
        transition: all 0.3s ease;
        cursor: pointer;
        height: 250px;
        display: flex;
        flex-direction: column;
        justify-content: center;
        align-items: center;
        position: relative;
        overflow: hidden;
        box-shadow: 0 5px 15px rgba(0,0,0,0.08);
    }
    
    .feature-card:hover {
        transform: translateY(-8px) scale(1.02);
        box-shadow: 0 15px 30px rgba(125, 60, 152, 0.15);
        border-color: #7D3C98;
    }
    
    .feature-icon {
        font-size: 3.5rem;
        margin-bottom: 1.5rem;
        background: linear-gradient(135deg, #7D3C98 0%, #8E44AD 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    
    .feature-title {
        font-size: 1.5rem;
        color: #2d3748;
        margin-bottom: 1rem;
        font-weight: 700;
    }
    
    .feature-description {
        font-size: 0.95rem;
        color: #4a5568;
        line-height: 1.5;
        font-weight: 400;
    }
    
    /* BOTÕES ROXOS */
    .stButton>button {
        background: linear-gradient(135deg, #7D3C98 0%, #8E44AD 100%) !important;
        color: white !important;
        border: none !important;
        padding: 0.7rem 1.5rem !important;
        border-radius: 10px !important;
        font-size: 0.95rem !important;
        font-weight: 600 !important;
        width: 100% !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 12px rgba(125, 60, 152, 0.3) !important;
    }
    
    .stButton>button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(125, 60, 152, 0.4) !important;
        background: linear-gradient(135deg, #8E44AD 0%, #9B59B6 100%) !important;
    }
    
    /* Botão voltar */
    .back-button {
        background: linear-gradient(135deg, #6c757d 0%, #495057 100%) !important;
        padding: 0.5rem 1rem !important;
        font-size: 0.85rem !important;
        margin-bottom: 1rem !important;
    }
    
    /* Barra de progresso roxa */
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #7D3C98 0%, #8E44AD 100%) !important;
    }
    
    .progress-text {
        font-weight: 600;
        color: #7D3C98;
        min-width: 60px;
        font-size: 1rem;
    }
    
    /* Cards de informação ROXOS */
    .info-card {
        background: linear-gradient(135deg, #7D3C98 0%, #8E44AD 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(125, 60, 152, 0.3);
    }
    
    .success-card {
        background: linear-gradient(135deg, #27ae60 0%, #2ecc71 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(39, 174, 96, 0.3);
    }
    
    .warning-card {
        background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
        color: white;
        padding: 1rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 5px 15px rgba(231, 76, 60, 0.3);
    }
    
    /* Seções expansíveis */
    .expandable-section {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
        border-left: 4px solid #7D3C98;
        box-shadow: 0 3px 10px rgba(0,0,0,0.08);
    }
    
    /* Templates */
    .template-card {
        background: white;
        border: 2px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.5rem;
        margin: 0.5rem 0;
        cursor: pointer;
        transition: all 0.3s ease;
        text-align: center;
        height: 120px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }
    
    .template-card:hover {
        border-color: #7D3C98;
        transform: translateY(-3px);
        box-shadow: 0 5px 15px rgba(125, 60, 152, 0.1);
    }
    
    .template-card.selected {
        border-color: #7D3C98;
        background: linear-gradient(135deg, rgba(125, 60, 152, 0.1) 0%, rgba(142, 68, 173, 0.1) 100%);
        box-shadow: 0 5px 15px rgba(125, 60, 152, 0.15);
    }
    
    /* Animações */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .fade-in {
        animation: fadeIn 0.4s ease-out;
    }
    
    /* Status indicators */
    .status-indicator {
        display: inline-block;
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 6px;
    }
    
    .status-processing {
        background: #f39c12;
        animation: pulse 1.5s infinite;
    }
    
    .status-completed {
        background: #27ae60;
    }
    
    @keyframes pulse {
        0% { opacity: 1; }
        50% { opacity: 0.5; }
        100% { opacity: 1; }
    }
    
    /* Textos sempre legíveis */
    .stMetric {
        color: #2d3748 !important;
    }
    
    .stMetric > div[data-testid="stMetricLabel"] > div {
        color: #4a5568 !important;
        font-weight: 600 !important;
        font-size: 0.9rem !important;
    }
    
    .stMetric > div[data-testid="stMetricValue"] > div {
        color: #2d3748 !important;
        font-weight: 700 !important;
        font-size: 1.1rem !important;
    }
    
    .stRadio > label {
        color: #2d3748 !important;
        font-weight: 500 !important;
    }
    
    .stNumberInput > label {
        color: #2d3748 !important;
        font-weight: 600 !important;
    }
    
    .stNumberInput > div > div > input {
        color: #2d3748 !important;
        background: white !important;
        border: 1px solid #e2e8f0 !important;
    }
    
    /* Textos gerais */
    .stMarkdown {
        color: #2d3748 !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        color: #2d3748 !important;
        margin-bottom: 0.75rem !important;
    }
    
    p, div, span {
        color: #2d3748 !important;
    }
    
    /* Ajustes para inputs */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stTextArea > div > div > textarea {
        background: white !important;
        color: #2d3748 !important;
        border: 1px solid #e2e8f0 !important;
    }
    
    /* Ajuste para os tabs */
    .stTabs [data-baseweb="tab-list"] {
        background: white !important;
        border-bottom: 1px solid #e2e8f0 !important;
    }
    
    .stTabs [data-baseweb="tab"] {
        background: white !important;
        color: #2d3748 !important;
    }
    
    .stTabs [aria-selected="true"] {
        background: white !important;
        color: #7D3C98 !important;
    }
    
    /* Forçar cores em todos os elementos */
    div[data-testid="stVerticalBlock"] {
        background: transparent !important;
    }
    
    section[data-testid="stFileUploadDropzone"] {
        background: rgba(255, 255, 255, 0.1) !important;
    }
</style>
""", unsafe_allow_html=True)

# NOVA FUNÇÃO: Copiar formatação entre worksheets
def copy_cell_formatting(source_cell, target_cell):
    """Copia a formatação de uma célula para outra"""
    try:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color
            )
        
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color
            )
        
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text
            )
        
        if source_cell.border:
            target_cell.border = Border(
                left=Side(style=source_cell.border.left.style, color=source_cell.border.left.color),
                right=Side(style=source_cell.border.right.style, color=source_cell.border.right.color),
                top=Side(style=source_cell.border.top.style, color=source_cell.border.top.color),
                bottom=Side(style=source_cell.border.bottom.style, color=source_cell.border.bottom.color)
            )
        
        target_cell.number_format = source_cell.number_format
        
    except Exception as e:
        st.warning(f"Aviso na formatação: {str(e)}")

# FUNÇÃO ATUALIZADA: Dividir planilhas mantendo formatação - SEMPRE COLAR A PARTIR DA LINHA 3
def split_spreadsheet_with_progress(file_path, progress_bar, progress_text, status_text, lines_per_file=None, num_files=None):
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        
        # Criar diretório temporário
        temp_dir = tempfile.mkdtemp()
        
        # Determinar se é CSV ou Excel
        is_csv = file_ext in ['.csv', '.txt']
        
        status_text.text("📖 Lendo arquivo...")
        progress_bar.progress(5)
        progress_text.text("5%")

        # Ler o arquivo completo
        if is_csv:
            # Para CSV, contar linhas primeiro
            with open(file_path, 'r', encoding='utf-8') as f:
                total_lines = sum(1 for line in f) - 1
            
            df = pd.read_csv(file_path, encoding='utf-8')
            original_workbook = None
        else:
            # Para Excel, carregar com openpyxl para manter formatação
            original_workbook = load_workbook(file_path)
            original_sheet = original_workbook.active
            
            # Ler dados com pandas também
            df = pd.read_excel(file_path, engine='openpyxl')
            total_lines = len(df)
        
        progress_bar.progress(15)
        progress_text.text("15%")
        status_text.text(f"📊 Arquivo lido: {total_lines} linhas encontradas")
        time.sleep(0.3)
        
        # Calcular linhas por arquivo
        if num_files:
            lines_per_file = ceil(total_lines / num_files)
        elif not lines_per_file:
            raise ValueError("Defina o número de linhas ou arquivos")
        
        total_parts = ceil(total_lines / lines_per_file)
        generated_files = []
        
        status_text.text("✂️ Iniciando divisão da planilha...")
        progress_bar.progress(20)
        progress_text.text("20%")
        
        for i in range(total_parts):
            start_idx = i * lines_per_file
            end_idx = min((i + 1) * lines_per_file, total_lines)
            
            # Atualizar progresso
            progress_percent = 20 + (i / total_parts) * 70
            progress_bar.progress(int(progress_percent))
            progress_text.text(f"{int(progress_percent)}%")
            status_text.text(f"📝 Criando parte {i+1} de {total_parts}...")
            
            # Para CSV, salvar normalmente
            if is_csv:
                part_df = df.iloc[start_idx:end_idx]
                output_file = os.path.join(temp_dir, f"{base_name}_parte_{i+1:02d}.csv")
                part_df.to_csv(output_file, index=False, encoding='utf-8')
            
            # Para Excel, manter formatação original - SEMPRE COLAR A PARTIR DA LINHA 3
            else:
                # Criar novo workbook
                new_wb = openpyxl.Workbook()
                new_ws = new_wb.active
                
                # CORREÇÃO: Copiar as DUAS PRIMEIRAS LINHAS com formatação (linhas 1 e 2)
                for row in range(1, 3):  # Linhas 1 e 2
                    if row <= original_sheet.max_row:
                        for col in range(1, original_sheet.max_column + 1):
                            source_cell = original_sheet.cell(row=row, column=col)
                            target_cell = new_ws.cell(row=row, column=col)
                            target_cell.value = source_cell.value
                            copy_cell_formatting(source_cell, target_cell)
                
                # CORREÇÃO: Copiar dados COM COLAR A PARTIR DA LINHA 3
                for row_idx in range(start_idx + 1, end_idx + 1):  # Ajustar índices
                    if row_idx + 1 <= original_sheet.max_row:  # +1 porque pulamos cabeçalho
                        for col in range(1, original_sheet.max_column + 1):
                            source_cell = original_sheet.cell(row=row_idx + 2, column=col)  # +2 para pular linhas 1 e 2
                            target_row = row_idx - start_idx + 2  # Começar na linha 3
                            target_cell = new_ws.cell(row=target_row, column=col)
                            target_cell.value = source_cell.value
                            copy_cell_formatting(source_cell, target_cell)
                
                # Ajustar largura das colunas
                for col in range(1, original_sheet.max_column + 1):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    new_ws.column_dimensions[column_letter].width = original_sheet.column_dimensions[column_letter].width
                
                output_file = os.path.join(temp_dir, f"{base_name}_parte_{i+1:02d}.xlsx")
                new_wb.save(output_file)
                new_wb.close()
            
            generated_files.append(output_file)
            time.sleep(0.05)
        
        # Fechar workbook original se existir
        if original_workbook:
            original_workbook.close()
        
        progress_bar.progress(95)
        progress_text.text("95%")
        status_text.text("📦 Preparando arquivos para download...")
        time.sleep(0.3)
        
        progress_bar.progress(100)
        progress_text.text("100%")
        status_text.text("✅ Processamento concluído!")
        
        return generated_files, temp_dir, total_parts
        
    except Exception as e:
        progress_bar.progress(0)
        progress_text.text("0%")
        status_text.text("❌ Erro no processamento")
        raise Exception(f"Erro ao dividir planilha: {str(e)}")

# CORREÇÃO: Função para carregar template e aplicar dados - ARQUIVOS SOLTOS - SEMPRE COLAR A PARTIR DA LINHA 3
def load_template_and_apply_data(uploaded_files, template_type):
    """Carrega o template e aplica os dados mantendo formatação - SEMPRE COLAR A PARTIR DA LINHA 3"""
    try:
        # CORREÇÃO: Template mapping com nomes exatos dos arquivos SOLTOS
        template_mapping = {
            "Clientes": "ClientesModeloExcel_Financeiro.xlsx",
            "Equipamentos": "EquipamentosModeloExcel.xlsx", 
            "Produtos": "ProdutosModeloExcel.xlsx",
            "Questionarios": "QuestionariosModeloExcel.xls"
        }
        
        template_filename = template_mapping.get(template_type)
        if not template_filename:
            raise ValueError(f"Template não encontrado para {template_type}")
        
        # CORREÇÃO: Buscar arquivo SOLTO (não na pasta modelos)
        template_path = template_filename
        
        if not os.path.exists(template_path):
            # Tentar buscar em diretório atual
            st.warning(f"⚠️ Template {template_filename} não encontrado no diretório atual. Verificando arquivos disponíveis...")
            
            # Listar arquivos disponíveis
            available_files = [f for f in os.listdir('.') if os.path.isfile(f)]
            st.write(f"Arquivos disponíveis: {available_files}")
            
            # Tentar encontrar arquivo similar
            matching_files = [f for f in available_files if template_type.lower() in f.lower()]
            if matching_files:
                template_path = matching_files[0]
                st.info(f"📂 Usando arquivo similar: {template_path}")
            else:
                raise FileNotFoundError(f"Arquivo template não encontrado: {template_filename}. Arquivos disponíveis: {available_files}")
        
        st.info(f"📂 Carregando template: {template_path}")
        
        # Carregar template com openpyxl para manter formatação
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active
        
        # Ler dados dos arquivos enviados
        all_data = []
        for uploaded_file in uploaded_files:
            file_ext = os.path.splitext(uploaded_file.name)[1].lower()
            
            if file_ext in ['.csv', '.txt']:
                df = pd.read_csv(uploaded_file, encoding='utf-8')
            else:
                df = pd.read_excel(uploaded_file, engine='openpyxl')
            
            all_data.append(df)
        
        # Combinar dados se houver múltiplos arquivos
        if len(all_data) > 1:
            combined_df = pd.concat(all_data, axis=0, ignore_index=True)
        else:
            combined_df = all_data[0]
        
        # Obter cabeçalhos do template (primeira linha)
        template_headers = []
        for col in range(1, template_ws.max_column + 1):
            cell_value = template_ws.cell(row=1, column=col).value
            template_headers.append(cell_value if cell_value else f"Coluna_{col}")
        
        st.info(f"📊 Estrutura do template: {len(template_headers)} colunas encontradas")
        
        # Mapear colunas dos dados para o template
        processed_data = []
        
        for _, row in combined_df.iterrows():
            new_row = {}
            for template_header in template_headers:
                # Tentar encontrar coluna correspondente nos dados
                if template_header in combined_df.columns:
                    new_row[template_header] = row[template_header]
                else:
                    # Verificar por correspondência aproximada
                    matching_cols = [col for col in combined_df.columns if str(col).lower() == str(template_header).lower()]
                    if matching_cols:
                        new_row[template_header] = row[matching_cols[0]]
                    else:
                        # Verificar por correspondência parcial
                        partial_matches = [col for col in combined_df.columns if str(template_header).lower() in str(col).lower() or str(col).lower() in str(template_header).lower()]
                        if partial_matches:
                            new_row[template_header] = row[partial_matches[0]]
                        else:
                            new_row[template_header] = ""  # Coluna vazia se não encontrar
            
            processed_data.append(new_row)
        
        # Criar DataFrame final com a estrutura do template
        final_df = pd.DataFrame(processed_data, columns=template_headers)
        
        st.success(f"✅ Dados processados: {len(final_df)} linhas mapeadas para o template")
        
        return final_df, template_headers, template_wb
        
    except Exception as e:
        raise Exception(f"Erro ao processar template: {str(e)}")

# CORREÇÃO: Função para juntar/formatar planilhas com template - SEMPRE COLAR A PARTIR DA LINHA 3
def merge_spreadsheets_with_template(uploaded_files, progress_bar, progress_text, status_text, template_type):
    try:
        status_text.text("🔍 Carregando template...")
        progress_bar.progress(10)
        progress_text.text("10%")
        
        # Carregar template e aplicar dados
        merged_df, template_columns, template_wb = load_template_and_apply_data(uploaded_files, template_type)
        
        progress_bar.progress(40)
        progress_text.text("40%")
        
        if len(uploaded_files) == 1:
            status_text.text("🎨 Aplicando formatação do template...")
        else:
            status_text.text("🔗 Combinando dados no template...")
        
        progress_bar.progress(70)
        progress_text.text("70%")
        
        # Criar arquivo final mantendo a formatação do template
        temp_dir = tempfile.mkdtemp()
        output_path = os.path.join(temp_dir, f"{template_type}_formatado.xlsx")
        
        # CORREÇÃO: Limpar dados existentes no template APENAS A PARTIR DA LINHA 3
        # Mantendo as duas primeiras linhas (1 e 2) intactas
        for row in range(3, template_wb.active.max_row + 1):  # Começar da linha 3
            for col in range(1, template_wb.active.max_column + 1):
                template_wb.active.cell(row=row, column=col).value = None
        
        # CORREÇÃO: Adicionar novos dados SEMPRE A PARTIR DA LINHA 3
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=False), 3):  # Começar na linha 3
            for c_idx, value in enumerate(row, 1):
                if r_idx <= template_wb.active.max_row and c_idx <= template_wb.active.max_column:
                    template_wb.active.cell(row=r_idx, column=c_idx).value = value
        
        # Salvar workbook
        template_wb.save(output_path)
        
        progress_bar.progress(95)
        progress_text.text("95%")
        status_text.text("💾 Salvando arquivo formatado...")
        time.sleep(0.3)
        
        progress_bar.progress(100)
        progress_text.text("100%")
        
        if len(uploaded_files) == 1:
            status_text.text("✅ Formatação concluída!")
        else:
            status_text.text("✅ Junção concluída!")
        
        return merged_df, template_columns, output_path
        
    except Exception as e:
        progress_bar.progress(0)
        progress_text.text("0%")
        status_text.text("❌ Erro no processamento")
        raise Exception(f"Erro ao processar planilhas: {str(e)}")

# Página inicial
def main_page():
    st.markdown("""
    <div class="main-container fade-in">
        <div class="main-header">Planilhas Setup Auvo</div>
        <div class="subtitle">Organize sua gestão de setups</div>
    """, unsafe_allow_html=True)
    
    # Cards principais
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">📊</div>
            <div class="feature-title">Dividir Planilhas</div>
            <div class="feature-description">
                Transforme planilhas extensas em arquivos menores e gerenciáveis. 
                <strong>Mantém a formatação original!</strong>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("🚀 Iniciar Divisão", key="split_btn", use_container_width=True):
            st.session_state.page = "split"
            st.rerun()

    with col2:
        st.markdown("""
        <div class="feature-card">
            <div class="feature-icon">🔗</div>
            <div class="feature-title">Formatar/Juntar Planilhas</div>
            <div class="feature-description">
                Formate planilhas individuais ou combine múltiplas planilhas seguindo templates padronizados.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("🚀 Iniciar Formatação", key="merge_btn", use_container_width=True):
            st.session_state.page = "merge"
            st.rerun()

# Página de divisão com progresso
def split_page():
    st.markdown("""
    <div class="main-container fade-in">
        <div class="main-header">Divisor de Planilhas</div>
        <div class="subtitle">Divida planilhas grandes em partes gerenciáveis <strong>mantendo a formatação original</strong></div>
    """, unsafe_allow_html=True)
    
    # Botão voltar
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("← Voltar", key="back_split", use_container_width=True, type="secondary"):
            st.session_state.page = "main"
            st.rerun()
    
    # Seção de upload
    with st.container():
        st.markdown('<div class="upload-section">', unsafe_allow_html=True)
        st.subheader("📁 Upload da Planilha")
        uploaded_file = st.file_uploader(
            "Selecione o arquivo para dividir",
            type=['xlsx', 'xls', 'csv', 'txt'],
            help="Arraste ou clique para selecionar o arquivo. Para Excel, a formatação será mantida!"
        )
        st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_file is not None:
        # Informações do arquivo
        with st.container():
            st.markdown('<div class="expandable-section compact-section">', unsafe_allow_html=True)
            st.subheader("📋 Informações do Arquivo")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Nome", uploaded_file.name)
            with col2:
                st.metric("Tamanho", f"{uploaded_file.size / 1024:.1f} KB")
            with col3:
                file_type = "Excel com Formatação" if uploaded_file.name.endswith(('.xlsx', '.xls')) else "CSV/Texto"
                st.metric("Tipo", file_type)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Configurações de divisão
        with st.container():
            st.markdown('<div class="expandable-section compact-section">', unsafe_allow_html=True)
            st.subheader("⚙️ Configurações de Divisão")
            
            col1, col2 = st.columns(2)
            with col1:
                split_method = st.radio(
                    "Método de Divisão:",
                    ["Dividir por número de linhas", "Dividir em número de arquivos"],
                    key="split_method"
                )
            
            with col2:
                if split_method == "Dividir por número de linhas":
                    lines_per_file = st.number_input(
                        "Linhas por arquivo:",
                        min_value=1,
                        value=5000,
                        help="Cada arquivo terá esta quantidade de linhas",
                        key="lines_input"
                    )
                    num_files = None
                else:
                    num_files = st.number_input(
                        "Número de arquivos:",
                        min_value=2,
                        value=5,
                        help="A planilha será dividida nesta quantidade de partes",
                        key="files_input"
                    )
                    lines_per_file = None
            
            # Informação sobre formatação
            if uploaded_file.name.endswith(('.xlsx', '.xls')):
                st.info("🎨 **Formatação preservada**: As planilhas Excel divididas manterão cores, fontes, bordas e formatação de células. **Dados sempre colados a partir da linha 3**.")
            else:
                st.info("📝 **Arquivo CSV**: A divisão será feita mantendo a estrutura de dados.")
                
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Área de processamento
        if st.button("🎯 Iniciar Processamento", type="primary", use_container_width=True):
            with st.container():
                st.markdown('<div class="expandable-section">', unsafe_allow_html=True)
                st.subheader("🔄 Processamento")
                
                # Barra de progresso com porcentagem
                progress_bar = st.progress(0)
                progress_col1, progress_col2 = st.columns([3, 1])
                with progress_col1:
                    status_text = st.empty()
                with progress_col2:
                    progress_text = st.empty()
                    progress_text.text("0%")
                
                try:
                    # Salvar arquivo temporariamente
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as tmp_file:
                        tmp_file.write(uploaded_file.getvalue())
                        tmp_path = tmp_file.name
                    
                    # Processar com barra de progresso
                    generated_files, temp_dir, total_parts = split_spreadsheet_with_progress(
                        tmp_path, progress_bar, progress_text, status_text, lines_per_file, num_files
                    )
                    
                    # Criar ZIP
                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
                        for file_path in generated_files:
                            zip_file.write(file_path, os.path.basename(file_path))
                    zip_buffer.seek(0)
                    
                    # Limpar arquivo temporário
                    os.unlink(tmp_path)
                    
                    # Resultados
                    st.markdown(f"""
                    <div class="success-card">
                        <h3 style="margin:0; color:white; font-size: 1.3rem;">✅ Processamento Concluído!</h3>
                        <p style="margin:0.5rem 0 0 0; color:white; font-size: 0.95rem;">
                        A planilha foi dividida em <strong>{total_parts}</strong> partes com sucesso.
                        {'<br>🎨 Formatação original preservada! Dados colados a partir da linha 3.' if uploaded_file.name.endswith(('.xlsx', '.xls')) else ''}
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Download
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 Baixar Todas as Partes (ZIP)",
                            data=zip_buffer,
                            file_name=f"{os.path.splitext(uploaded_file.name)[0]}_partes.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                    
                    # Preview
                    st.subheader("👀 Preview dos Resultados")
                    tabs = st.tabs([f"Parte {i+1}" for i in range(min(3, len(generated_files)))])
                    
                    for i, tab in enumerate(tabs):
                        with tab:
                            if i < len(generated_files):
                                file_path = generated_files[i]
                                file_ext = os.path.splitext(file_path)[1].lower()
                                
                                try:
                                    if file_ext == '.csv':
                                        df_preview = pd.read_csv(file_path)
                                    else:
                                        df_preview = pd.read_excel(file_path, engine='openpyxl')
                                    
                                    st.write(f"**Parte {i+1}** - {len(df_preview)} linhas × {len(df_preview.columns)} colunas")
                                    st.dataframe(df_preview.head(6), use_container_width=True)
                                    
                                except Exception as e:
                                    st.error(f"Erro ao visualizar: {str(e)}")
                    
                    # Limpeza
                    for file_path in generated_files:
                        if os.path.exists(file_path):
                            os.unlink(file_path)
                    if os.path.exists(temp_dir):
                        os.rmdir(temp_dir)
                        
                except Exception as e:
                    st.error(f"❌ Erro no processamento: {str(e)}")
                
                st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)

# Página de junção com templates
def merge_page():
    st.markdown("""
    <div class="main-container fade-in">
        <div class="main-header">Formatar e Juntar Planilhas</div>
        <div class="subtitle">Padronize planilhas individuais ou combine múltiplas planilhas</div>
    """, unsafe_allow_html=True)
    
    # Botão voltar
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("← Voltar", key="back_merge", use_container_width=True, type="secondary"):
            st.session_state.page = "main"
            st.rerun()
    
    # Seleção de template
    with st.container():
        st.markdown('<div class="expandable-section compact-section">', unsafe_allow_html=True)
        st.subheader("🎯 Selecione o Tipo de Planilha")
        
        template_options = {
            "Clientes": "Template para dados de clientes",
            "Equipamentos": "Template para cadastro de equipamentos", 
            "Produtos": "Template para catálogo de produtos",
            "Questionarios": "Template para formulários de questionários"
        }
        
        cols = st.columns(4)
        selected_template = st.session_state.get('selected_template', None)
        
        for i, (template_name, description) in enumerate(template_options.items()):
            with cols[i]:
                is_selected = selected_template == template_name
                css_class = "template-card selected" if is_selected else "template-card"
                
                st.markdown(f"""
                <div class="{css_class}">
                    <h4 style="margin:0; font-size: 1.1rem;">{template_name}</h4>
                    <p style="margin:0.5rem 0 0 0; font-size: 0.8rem; color: #6c757d;">{description}</p>
                </div>
                """, unsafe_allow_html=True)
                
                if st.button("Selecionar", key=f"template_{i}", use_container_width=True):
                    st.session_state.selected_template = template_name
                    st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)
    
    if st.session_state.get('selected_template'):
        selected_template = st.session_state.selected_template
        
        # Upload de arquivos
        with st.container():
            st.markdown('<div class="upload-section">', unsafe_allow_html=True)
            st.subheader(f"📁 Upload das Planilhas de {selected_template}")
            
            uploaded_files = st.file_uploader(
                f"Selecione as planilhas de {selected_template} para formatar/juntar",
                type=['xlsx', 'xls', 'csv', 'txt'],
                accept_multiple_files=True,
                help="Selecione uma ou mais planilhas do mesmo tipo"
            )
            
            if uploaded_files:
                st.write(f"**Arquivos selecionados:** {len(uploaded_files)}")
                
                for i, file in enumerate(uploaded_files):
                    col1, col2 = st.columns([3, 1])
                    with col1:
                        st.write(f"**{i+1}. {file.name}**")
                    with col2:
                        st.write(f"`{file.size / 1024:.1f} KB`")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Processamento
        if uploaded_files:
            button_label = "🎯 Iniciar Formatação" if len(uploaded_files) == 1 else "🎯 Iniciar Junção"
            
            if st.button(button_label, type="primary", use_container_width=True):
                with st.container():
                    st.markdown('<div class="expandable-section">', unsafe_allow_html=True)
                    st.subheader("🔄 Processamento")
                    
                    # Barra de progresso com porcentagem
                    progress_bar = st.progress(0)
                    progress_col1, progress_col2 = st.columns([3, 1])
                    with progress_col1:
                        status_text = st.empty()
                    with progress_col2:
                        progress_text = st.empty()
                        progress_text.text("0%")
                    
                    try:
                        merged_df, template_columns, output_path = merge_spreadsheets_with_template(
                            uploaded_files, progress_bar, progress_text, status_text, selected_template
                        )
                        
                        # Ler arquivo final para download
                        with open(output_path, 'rb') as f:
                            excel_data = f.read()
                        
                        # Mensagem dinâmica
                        if len(uploaded_files) == 1:
                            success_message = f"""
                            <div class="success-card">
                                <h3 style="margin:0; color:white; font-size: 1.3rem;">✅ Formatação Concluída!</h3>
                                <p style="margin:0.5rem 0 0 0; color:white; font-size: 0.95rem;">
                                Planilha formatada com sucesso no template: <strong>{len(merged_df)}</strong> linhas e <strong>{len(merged_df.columns)}</strong> colunas.
                                <br><strong>Dados colados a partir da linha 3</strong> - Linhas 1 e 2 mantidas fixas.
                                </p>
                            </div>
                            """
                        else:
                            success_message = f"""
                            <div class="success-card">
                                <h3 style="margin:0; color:white; font-size: 1.3rem;">✅ Junção Concluída!</h3>
                                <p style="margin:0.5rem 0 0 0; color:white; font-size: 0.95rem;">
                                <strong>{len(uploaded_files)}</strong> planilhas combinadas no template: <strong>{len(merged_df)}</strong> linhas e <strong>{len(merged_df.columns)}</strong> colunas.
                                <br><strong>Dados colados a partir da linha 3</strong> - Linhas 1 e 2 mantidas fixas.
                                </p>
                            </div>
                            """
                        
                        st.markdown(success_message, unsafe_allow_html=True)
                        
                        # Preview
                        st.subheader("👀 Preview do Resultado")
                        st.dataframe(merged_df.head(10), use_container_width=True)
                        
                        # Download
                        st.subheader("📥 Download do Arquivo Formatado")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.download_button(
                                label="📥 Excel com Formatação",
                                data=excel_data,
                                file_name=f"{selected_template.lower()}_formatado.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        
                        with col2:
                            csv_buffer = BytesIO()
                            csv_buffer.write(merged_df.to_csv(index=False, encoding='utf-8').encode('utf-8'))
                            csv_buffer.seek(0)
                            
                            st.download_button(
                                label="📥 CSV",
                                data=csv_buffer,
                                file_name=f"{selected_template.lower()}_formatado.csv",
                                mime="text/csv",
                                use_container_width=True
                            )
                        
                        # Limpeza
                        if os.path.exists(output_path):
                            os.unlink(output_path)
                        if os.path.exists(os.path.dirname(output_path)):
                            shutil.rmtree(os.path.dirname(output_path))
                            
                    except Exception as e:
                        st.error(f"❌ Erro no processamento: {str(e)}")
                    
                    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("</div>", unsafe_allow_html=True)

# Gerenciamento de estado
if 'page' not in st.session_state:
    st.session_state.page = 'main'

if 'selected_template' not in st.session_state:
    st.session_state.selected_template = None

# Navegação 
if st.session_state.page == 'main':
    main_page()
elif st.session_state.page == 'split':
    split_page()
elif st.session_state.page == 'merge':
    merge_page()
